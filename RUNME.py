
"""
Timeoversigt_v2.4 LTS — VISUEL+ — DK RUNME (med indlejret Python)

Formål
• Generér en komplet Timeoversigt (Excel) ud fra en Portfolio Controlling-CSV.
• Faner: Sheet1, SEKTIONSOPSUMMERING, DATA ISSUES.
• Al visuel formatering (farver, talformater, kolonneorden, fjernelse af 'Time stamp') styres deterministisk i Python.

Sådan bruges RUNME
1) Upload i samme besked:
   • Denne RUNME-fil (.txt)
   • Portfolio Controlling-CSV’en
2) Skriv i samme besked (copy/paste):
   Kør Python-koden i denne promptfil på denne CSV-fil. Generér den færdige Timeoversigt-Excel.
3) Copilot kører Python-koden herunder og leverer en Excel med VISUEL+ og danske talformater.

Vigtige ændringer i v2.4
• Tekstkolonner låses i Excel (number_format='@'): Sektionsnr., Projekt org., Projekt nr., UK, Opgave nr., Year; samt Sektionsnr. i SEKTIONSOPSUMMERING.
• DATA ISSUES: Forecast og Actuals vises som beløb med tusindtalsseparator og 2 decimaler (dansk: #.##0,00).
• UK-mapping (Option A): Kun UK11 → UK10 indgår i UK10-buckets. Andre UK ≠ 10/90/95/97 registreres som issues, men indgår ikke i UK-buckets.
• Alt VISUEL+ bevares.
"""


import pandas as pd, numpy as np, os, calendar
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# VISUEL+ farver
DETAIL_CALC_BLUE = "FFEAF2FF"
TOTAL_RAW_GRAY   = "FFEFEFEF"
TOTAL_CALC_BLUE  = "FFCFE5FF"
BANNER_FILL      = "FFDDEBF7"
GREEN="FFC6EFCE"; YELLOW="FFFFEB9C"; RED="FFFFC7CE"

# Kolonne-grupper
RAW_COLS=['Institut','Sektionsnr.','Medarbejder','Ans. Start','Ans. Slut','Nuv. normtid','Projekt org.','Projekt nr.','Projekt navn','UK','Opgave nr.','Opgave navn','Year','Forecast','Actuals','Difference']
CALC_COLS=['Ans_grad_decimal','Maaneder_ansat_2026','Normtid_2026','Planlaegningsgrad','UK10_timer','UK90_timer','UK95_timer','UK97_timer','UK95_97_timer','UK10_pct','UK90_pct','UK95_pct','UK97_pct','UK95_97_pct']
PCT_COLS_SHEET1=['Planlaegningsgrad','UK10_pct','UK90_pct','UK95_pct','UK97_pct','UK95_97_pct']
SHEET1_ORDER = RAW_COLS + CALC_COLS + ['UK10_contains_mapped']

SEK_ORDER=['Sektionsnr.','antal_medarbejdere','total_forecast','total_normtid','Planlaegningsgrad_faktisk','UK10_timer','UK90_timer','UK95_timer','UK97_timer','UK95_97_timer','UK10_pct','UK90_pct','UK95_pct','UK97_pct','UK95_97_pct']
PCT_COLS_SEK=['Planlaegningsgrad_faktisk','UK10_pct','UK90_pct','UK95_pct','UK97_pct','UK95_97_pct']

# Kolonner der skal låses som TEKST i Excel
TEXT_COLS_SHEET1 = ['Sektionsnr.','Projekt org.','Projekt nr.','UK','Opgave nr.','Year']
TEXT_COLS_SEK = ['Sektionsnr.']

PLAN_GREEN=(0.98,1.01); PLAN_YELLOW_LOW=(0.95,0.98); PLAN_YELLOW_HIGH=(1.01,1.05)

# Helpers
from datetime import datetime as dt

def parse_number(x):
    if pd.isna(x): return 0.0
    s=str(x).strip()
    if s=="": return 0.0
    s=s.replace(" ","")
    if "," in s: s=s.replace(".","").replace(",",".")
    try: return float(s)
    except:
        try: return float("".join(ch for ch in s if (ch.isdigit() or ch in "+-.")))
        except: return 0.0

def parse_date(val):
    if pd.isna(val): return None
    try: return pd.to_datetime(val, dayfirst=True).date()
    except: return None

def months_in_2026(s,e):
    from datetime import date as _d
    year=2026; s=s or _d(1900,1,1); e=e or _d(2099,12,31)
    if e<_d(year,1,1) or s>_d(year,12,31): return 0
    import calendar
    c=0
    for m in range(1,13):
        ms=_d(year,m,1); me=_d(year,m,calendar.monthrange(year,m)[1])
        if not(e<ms or s>me): c+=1
    return c

def normtid(g,m): return float(g)*1591.0*(m/12.0)

DK_NUMBER='#,##0.00'
DK_PERCENT='0.00%'


def write_excel(path, sheet1, sek, issues, banner):
    wb=Workbook()

    # SHEET1
    ws=wb.active; ws.title='Sheet1'
    ws.append([banner]); ws.cell(1,1).fill=PatternFill(start_color=BANNER_FILL,end_color=BANNER_FILL,fill_type='solid')
    ws.append(list(sheet1.columns))
    for j in range(1, ws.max_column+1): ws.cell(2,j).font=Font(bold=True)
    for _,r in sheet1.iterrows(): ws.append(list(r))
    hdr={ws.cell(2,i).value:i for i in range(1, ws.max_column+1)}
    med_idx=hdr.get('Medarbejder')

    for rr in range(3, ws.max_row+1):
        mv = ws.cell(rr, med_idx).value if med_idx else None
        is_total = isinstance(mv, str) and mv.startswith('Total for ')
        # Farvelag
        for c in RAW_COLS:
            if c in hdr:
                cell=ws.cell(rr,hdr[c])
                cell.fill=PatternFill(start_color=TOTAL_RAW_GRAY,end_color=TOTAL_RAW_GRAY,fill_type='solid') if is_total else PatternFill(fill_type=None)
        for c in CALC_COLS:
            if c in hdr:
                base = TOTAL_CALC_BLUE if is_total else DETAIL_CALC_BLUE
                ws.cell(rr,hdr[c]).fill=PatternFill(start_color=base,end_color=base,fill_type='solid')
        # Procenter
        for c in PCT_COLS_SHEET1:
            if c in hdr: ws.cell(rr,hdr[c]).number_format=DK_PERCENT
        # Beløb (tal)
        numeric_cols = set(RAW_COLS + CALC_COLS) - set(PCT_COLS_SHEET1)
        for c in numeric_cols:
            if c in hdr:
                cell=ws.cell(rr,hdr[c])
                if isinstance(cell.value,(int,float)): cell.number_format=DK_NUMBER
        # Planlægningsgrad farvelægning på totalrækker
        if is_total and 'Planlaegningsgrad' in hdr:
            v=ws.cell(rr,hdr['Planlaegningsgrad']).value
            try:
                if PLAN_GREEN[0]<=v<=PLAN_GREEN[1]: col=GREEN
                elif (PLAN_YELLOW_LOW[0]<=v<PLAN_YELLOW_LOW[1]) or (PLAN_YELLOW_HIGH[0]<v<=PLAN_YELLOW_HIGH[1]): col=YELLOW
                else: col=RED
                ws.cell(rr,hdr['Planlaegningsgrad']).fill=PatternFill(start_color=col,end_color=col,fill_type='solid')
            except: pass
        # Lås tekstkolonner som tekst (Excel må ikke omformatere)
        for c in TEXT_COLS_SHEET1:
            if c in hdr:
                ws.cell(rr, hdr[c]).number_format='@'

    # UK10 gul markering når der er UK11-mapping
    flag_idx=hdr.get('UK10_contains_mapped'); uk10_t=hdr.get('UK10_timer'); uk10_p=hdr.get('UK10_pct')
    if flag_idx and uk10_t and uk10_p:
        from openpyxl.utils import get_column_letter as _g
        ws.column_dimensions[_g(flag_idx)].hidden=True
        for rr in range(3, ws.max_row+1):
            mv = ws.cell(rr, med_idx).value if med_idx else None
            if isinstance(mv,str) and mv.startswith('Total for '):
                flag = ws.cell(rr, flag_idx).value
                if str(flag).lower() in ('true','1','yes'):
                    ws.cell(rr, uk10_t).fill=PatternFill(start_color=YELLOW,end_color=YELLOW,fill_type='solid')
                    ws.cell(rr, uk10_p).fill=PatternFill(start_color=YELLOW,end_color=YELLOW,fill_type='solid')

    # Autosize
    for col in ws.columns:
        l = get_column_letter(col[0].column)
        ml = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[l].width = min(max(10, ml+2), 60)

    # SEKTIONSOPSUMMERING
    w2=wb.create_sheet('SEKTIONSOPSUMMERING')
    w2.append([banner]); w2.cell(1,1).fill=PatternFill(start_color=BANNER_FILL,end_color=BANNER_FILL,fill_type='solid')
    w2.append(SEK_ORDER)
    for j in range(1,len(SEK_ORDER)+1): w2.cell(2,j).font=Font(bold=True)
    for _,r in sek.iterrows(): w2.append([r.get(c) for c in SEK_ORDER])
    hdr2={w2.cell(2,i).value:i for i in range(1, w2.max_column+1)}
    for rr in range(3, w2.max_row+1):
        for c in PCT_COLS_SEK:
            i=hdr2[c]; w2.cell(rr,i).number_format=DK_PERCENT
        for c in SEK_ORDER:
            if c not in PCT_COLS_SEK and c!='antal_medarbejdere':
                i=hdr2[c]; cell=w2.cell(rr,i)
                if isinstance(cell.value,(int,float)): cell.number_format=DK_NUMBER
        # Lås tekstkolonner (kun Sektionsnr.)
        for c in TEXT_COLS_SEK:
            if c in hdr2:
                w2.cell(rr, hdr2[c]).number_format='@'
        # Farvelæg planlægningsgrad
        pi=hdr2['Planlaegningsgrad_faktisk']; pv=w2.cell(rr,pi).value
        try:
            if PLAN_GREEN[0]<=pv<=PLAN_GREEN[1]: col=GREEN
            elif (PLAN_YELLOW_LOW[0]<=pv<PLAN_YELLOW_LOW[1]) or (PLAN_YELLOW_HIGH[0]<pv<=PLAN_YELLOW_HIGH[1]): col=YELLOW
            else: col=RED
            w2.cell(rr,pi).fill=PatternFill(start_color=col,end_color=col,fill_type='solid')
        except: pass

    # Autosize
    for col in w2.columns:
        l = get_column_letter(col[0].column)
        ml = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        w2.column_dimensions[l].width = min(max(10, ml+2), 60)

    # DATA ISSUES
    w3=wb.create_sheet('DATA ISSUES')
    w3.append([banner]); w3.cell(1,1).fill=PatternFill(start_color=BANNER_FILL,end_color=BANNER_FILL,fill_type='solid')
    if not issues.empty:
        w3.append(list(issues.columns))
        for j in range(1,len(issues.columns)+1): w3.cell(2,j).font=Font(bold=True)
        for _,r in issues.iterrows(): w3.append([r.get(c) for c in issues.columns])
        # Formater beløbskolonner
        hdr3={w3.cell(2,i).value:i for i in range(1, w3.max_column+1)}
        for rr in range(3, w3.max_row+1):
            for colname in ['Forecast','Actuals']:
                if colname in hdr3:
                    cell=w3.cell(rr, hdr3[colname])
                    if isinstance(cell.value,(int,float)): cell.number_format=DK_NUMBER

    # Autosize
    for col in w3.columns:
        l = get_column_letter(col[0].column)
        ml = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        w3.column_dimensions[l].width = min(max(10, ml+2), 60)

    wb.save(path)


def run_pipeline(csv_path, out_path, institute_hint=None):
    df=pd.read_csv(csv_path,sep=';')
    df.columns=[c.strip() for c in df.columns]
    if 'Year' in df.columns: df=df[df['Year'].astype(str).str.contains('2026')]

    # Beløb til tal
    for c in ['Forecast','Actuals','Difference']:
        if c in df.columns: df[c]=df[c].apply(parse_number)
        else: df[c]=0.0

    # Fjern 'Time stamp'
    if 'Time stamp' in df.columns: df=df.drop(columns=['Time stamp'])

    # Ansættelsesmeta per medarbejder
    empmeta={}
    for _,r in df.iterrows():
        e=r.get('Medarbejder')
        if pd.isna(e): continue
        if e not in empmeta:
            s=parse_date(r.get('Ans. Start')); en=parse_date(r.get('Ans. Slut')); g=parse_number(r.get('Nuv. normtid'))
            m=months_in_2026(s,en)
            empmeta[e]={'Ans_grad_decimal': g if g>0 else 0.0,'Maaneder_ansat_2026': m,'Normtid_2026': normtid(g,m)}
    for key in ['Ans_grad_decimal','Maaneder_ansat_2026','Normtid_2026']:
        df[key]=df['Medarbejder'].map(lambda x: empmeta.get(x,{}).get(key, 0 if key!='Normtid_2026' else 0.0))

    allowed={10,90,95,97}
    totals=[]; issues_rows=[]

    for e,g in df.groupby('Medarbejder', dropna=True):
        f=g['Forecast'].sum(); a=g['Actuals'].sum(); d=g['Difference'].sum()
        uk10=g.loc[g['UK']==10,'Forecast'].sum()
        uk11_map=g.loc[g['UK']==11,'Forecast'].sum()  # Option A mapping
        uk90=g.loc[g['UK']==90,'Forecast'].sum(); uk95=g.loc[g['UK']==95,'Forecast'].sum(); uk97=g.loc[g['UK']==97,'Forecast'].sum()

        other_forbid_mask=(~g['UK'].isin(list(allowed))) & (g['UK']!=11)
        has_mapped = abs(uk11_map)>1e-12

        norm=empmeta.get(e,{}).get('Normtid_2026',0.0)
        plan=(f/norm) if norm and abs(norm)>1e-12 else np.nan

        totals.append({
            'Institut': g['Institut'].iloc[0] if 'Institut' in g.columns and len(g)>0 else institute_hint,
            'Sektionsnr.': g['Sektionsnr.'].iloc[0] if 'Sektionsnr.' in g.columns and len(g)>0 else None,
            'Medarbejder': f'Total for {e}',
            'Ans. Start': None,'Ans. Slut': None,'Nuv. normtid': empmeta.get(e,{}).get('Ans_grad_decimal',0.0),
            'Projekt org.': None,'Projekt nr.': None,'Projekt navn': None,'UK': None,'Opgave nr.': None,'Opgave navn': None,
            'Year': 2026,
            'Forecast': round(f,6),'Actuals': round(a,6),'Difference': round(d,6),
            'Ans_grad_decimal': empmeta.get(e,{}).get('Ans_grad_decimal',0.0),
            'Maaneder_ansat_2026': empmeta.get(e,{}).get('Maaneder_ansat_2026',0),
            'Normtid_2026': round(norm,6), 'Planlaegningsgrad': plan,
            'UK10_timer': round(uk10+uk11_map,6),'UK90_timer': round(uk90,6),'UK95_timer': round(uk95,6),'UK97_timer': round(uk97,6),
            'UK95_97_timer': round(uk95+uk97,6),
            'UK10_pct': (uk10+uk11_map)/f if f else np.nan,
            'UK90_pct': uk90/f if f else np.nan,
            'UK95_pct': uk95/f if f else np.nan,
            'UK97_pct': uk97/f if f else np.nan,
            'UK95_97_pct': (uk95+uk97)/f if f else np.nan,
            'UK10_contains_mapped': has_mapped
        })

        # Issues for UK11 (mapped) og øvrige forbudte UK
        for _, rr in g[g['UK']==11].iterrows():
            if abs(rr.get('Forecast',0.0))!=0:
                issues_rows.append({
                    'Medarbejder': e,
                    'Sektionsnr.': rr.get('Sektionsnr.'),
                    'Projekt nr.': rr.get('Projekt nr.'),
                    'Opgave nr.': rr.get('Opgave nr.'),
                    'UK': rr.get('UK'),
                    'Forecast': rr.get('Forecast'),
                    'Actuals': rr.get('Actuals'),
                    'Fejlkategori': 'UK11 mapped til UK10',
                    'Beskrivelse': 'UK11 indgår i UK10-buckets (Option A).',
                    'Forslag til handling': 'Overvej om UK11 bør flyttes til UK10 i kildedata.'
                })
        if other_forbid_mask.any():
            for _, rr in g[other_forbid_mask].iterrows():
                if abs(rr.get('Forecast',0.0))!=0:
                    issues_rows.append({
                        'Medarbejder': e,
                        'Sektionsnr.': rr.get('Sektionsnr.'),
                        'Projekt nr.': rr.get('Projekt nr.'),
                        'Opgave nr.': rr.get('Opgave nr.'),
                        'UK': rr.get('UK'),
                        'Forecast': rr.get('Forecast'),
                        'Actuals': rr.get('Actuals'),
                        'Fejlkategori': 'Forbudt UK-kode',
                        'Beskrivelse': 'UK≠10/90/95/97 (og ikke UK11). Indgår ikke i UK-buckets.',
                        'Forslag til handling': 'Flyt til godkendt UK i Portfoliostyring.'
                    })
        if (norm==0) or (abs(norm)<1e-12):
            issues_rows.append({
                'Medarbejder': e,
                'Sektionsnr.': g['Sektionsnr.'].iloc[0] if 'Sektionsnr.' in g.columns else None,
                'Projekt nr.': None,'Opgave nr.': None,'UK': None,
                'Forecast': f,'Actuals': a,
                'Fejlkategori': 'Normtid_2026 = 0',
                'Beskrivelse': 'Planlægningsgrad kan ikke beregnes (division med 0).',
                'Forslag til handling': 'Kontrollér ansættelsesgrad og ansættelsesperioder.'
            })

    # Sheet1 med detaljer + totalrækker
    out_rows=[]
    for e,g in df.groupby('Medarbejder', dropna=True, sort=False):
        out_rows.append(g)
        tot=[t for t in totals if t['Medarbejder']==f'Total for {e}']
        if tot: out_rows.append(pd.DataFrame([tot[0]]))
    sheet1=pd.concat(out_rows, ignore_index=True) if out_rows else df.copy()
    for c in SHEET1_ORDER:
        if c not in sheet1.columns: sheet1[c]=np.nan
    sheet1=sheet1[SHEET1_ORDER]

    # Sektionsopsummering fra totals
    totals_df=pd.DataFrame(totals)
    if totals_df.empty:
        sek=pd.DataFrame(columns=SEK_ORDER)
    else:
        g=totals_df.groupby('Sektionsnr.', dropna=False)
        agg=g.agg(antal_medarbejdere=('Medarbejder','count'), total_forecast=('Forecast','sum'), total_normtid=('Normtid_2026','sum'), UK10_timer=('UK10_timer','sum'), UK90_timer=('UK90_timer','sum'), UK95_timer=('UK95_timer','sum'), UK97_timer=('UK97_timer','sum')).reset_index()
        agg['UK95_97_timer']=agg['UK95_timer']+agg['UK97_timer']
        def pct2(x, tot): return (x/tot) if abs(tot)>1e-12 else np.nan
        agg['Planlaegningsgrad_faktisk']=agg.apply(lambda r: (r['total_forecast']/r['total_normtid']) if abs(r['total_normtid'])>1e-12 else np.nan, axis=1)
        agg['UK10_pct']=agg.apply(lambda r: pct2(r['UK10_timer'], r['total_forecast']), axis=1)
        agg['UK90_pct']=agg.apply(lambda r: pct2(r['UK90_timer'], r['total_forecast']), axis=1)
        agg['UK95_pct']=agg.apply(lambda r: pct2(r['UK95_timer'], r['total_forecast']), axis=1)
        agg['UK97_pct']=agg.apply(lambda r: pct2(r['UK97_timer'], r['total_forecast']), axis=1)
        agg['UK95_97_pct']=agg.apply(lambda r: pct2(r['UK95_97_timer'], r['total_forecast']), axis=1)
        for c in SEK_ORDER:
            if c not in agg.columns: agg[c]=np.nan
        sek=agg[SEK_ORDER]

    banner=f"Timeoversigt v2.4 LTS — VISUEL+ — DK — input: {os.path.basename(csv_path)} — genereret: {datetime.now():%Y-%m-%d %H:%M}"
    write_excel(out_path, sheet1, sek, pd.DataFrame(issues_rows), banner)
